''' used to replace borked shipping schedule
'''
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import Border, Side
from datetime import datetime
from pathlib import Path

#filepath = Path('/mnt/f/SCHEDULES/2020 schedules/Shipping Schedule 2020.xlsx')
filepath = Path('schedule.xlsx')
assert filepath.is_file()

wb = load_workbook(filename=filepath, read_only=True)

print('sheetnames', wb.sheetnames)

out_filepath = Path('fixed.xlsx')
if out_filepath.is_file():
    out_filepath.unlink()

new_workbook = Workbook()

top_border = Border(top=Side(style='thin'))
border = Border(top=Side(style='thin'),bottom=Side(style='thin'))
bottom_border = Border(bottom=Side(style='thin'))

def stringify(v):
    t = type(v)
    if t is datetime:
        return v.date()
    return v

widths = [12, 12, 4, 8, 8, 4, 30, 30, 20, 20, 16, 16, 4, 16, 16, 16, 30]
for sheet in wb:
    new_sheet = new_workbook.create_sheet(title=sheet.title)
    done = False

    for i, row in enumerate(sheet.iter_rows()):
        vals = [stringify(c.value) for c in row]
        new_sheet.append(vals)

        if i > 2 and not done:
            pcell = sheet.cell(row=i, column=1).value
            ccell = sheet.cell(row=i+1, column=1).value
            ncell = sheet.cell(row=i+2, column=1).value

            if ccell not in [None, '']:
                new_sheet.cell(row=i+1, column=15).value = '={0}{2}*{1}{2}'.format(*map(get_column_letter, [14, 4]), i+1)
                new_sheet.cell(row=i+1, column=16).value = '={0}{2}*{1}{2}'.format(*map(get_column_letter, [14, 5]), i+1)
                for j in range(14,18):
                    new_sheet.cell(row=i+1, column=j).number_format = '0.00'



            if ccell in [None, ''] and ncell in [None, '']:
                done = True
            elif ccell in [None, '']:
                pass
            elif pcell in [None, ''] and ncell in [None, '']:
                for cell in new_sheet[i + 1][:16]:
                    cell.border = border
            elif ncell in [None, '']:
                for cell in new_sheet[i + 1][:16]:
                    cell.border = bottom_border
            elif pcell in [None, '']:
                for cell in new_sheet[i + 1][:16]:
                    cell.border = top_border

    for cell in new_sheet[2][:16]:
        cell.border = Border(bottom=Side(style='thick'))


    for i, w in enumerate(widths):
        new_sheet.column_dimensions[get_column_letter(i + 1)].width = w

del new_workbook['Sheet']

#for column worksheet.column_dimensions[get_column_letter(i+1)].width
new_workbook.save(out_filepath)
